import datetime

import openpyxl
import  xlrd
import random
import copy
import docx
from docx import Document
from openpyxl.styles import Border,Side,Font


class Job:
    #日计划，月计划，等级， 线路，施工名称，施工日期，上下行，施工时间，登记地点
    daily_id = ""
    monthly_id = ""
    level = ""
    railway = ""
    jobName = ""
    jobDate = ""
    jobPlace = ""
    job_upOrDown = ""
    jobTime = ""
    job_registrationPlace = ""
    pass

#所有文字
allText = ""
#所有施工
allJobData = []

def getText():
    global allText
    global allJobData
    document = Document("计划.docx")
    all_paragraphs = document.paragraphs
    text = ""
    for paragraph in all_paragraphs:
        text = text +"|"+ paragraph.text.replace("：",":").replace(" ",":")
        print(paragraph.text)
    #找到一组就存一下
    loading = False
    hasGotOne = False
    currentJob = Job()
    for str in text.split("|"):
        if "日计划:" in  str:
            if loading == False:
                loading = True
            # 找到第二个或者多个了，前一个存起来
            elif loading == True:
                allJobData.append(currentJob)
                currentJob = Job()
            currentJob.daily_id = str.replace("日计划", "").replace(":", "")
            allText = allText + str
        elif "月计划:" in str:
            currentJob.monthly_id = str.replace("月计划", "").replace(":", "")
        elif "等级:" in str:
            currentJob.level = str.replace("等级", "").replace(":", "")
        elif "线路:" in str:
            currentJob.railway = str.replace("线路", "").replace(":", "")
        elif "施工项目:" in str:
            currentJob.jobName = str.replace("施工项目", "").replace(":", "")
        elif "施工日期:" in str:
            currentJob.jobDate = str.replace("施工日期", "").replace(":", "")
        elif "施工地点:" in str:
            currentJob.jobPlace = str.replace("施工地点", "").replace(":", "")
        elif "施工行别:" in str:
            currentJob.job_upOrDown = str.replace("施工行别", "").replace(":", "")
        elif "施工时间:" in str:
            currentJob.jobTime = str.replace("施工时间:", "")
        elif "施工内容及影响范围:" in str:
            currentJob.job_registrationPlace = str.replace("施工内容及影响范围:", "").replace(":", "")
    allJobData.append(currentJob)
    return text
    pass

def fillExcel():
    global allJobData
    target_workbook = openpyxl.load_workbook("模板.xlsx")
    work_sheet = target_workbook.worksheets[0]
    #字体
    font = Font(name="宋体", size=10, bold=False, italic=False)
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
    #记录一下一二三级施工的个数
    allLevels = ""
    count = 0
    #从第三行开始写
    for row_number in range(3,len(allJobData)+3):
        temp_data = allJobData[row_number-3]
        work_sheet.cell(row_number, 1).value = row_number - 2
        work_sheet.cell(row_number, 2).value = temp_data.daily_id
        work_sheet.cell(row_number, 3).value = temp_data.level
        allLevels = allLevels + "-" + temp_data.level
        work_sheet.cell(row_number, 4).value = temp_data.railway
        work_sheet.cell(row_number, 5).value = temp_data.jobName
        work_sheet.cell(row_number, 6).value = temp_data.jobDate
        work_sheet.cell(row_number, 7).value = temp_data.jobPlace
        work_sheet.cell(row_number, 8).value = temp_data.jobTime
        row_number = row_number + 1
        count = row_number
    work_sheet.cell(count, 1).value = "注：1.高铁队辖区I级施工"+str(allLevels.count("Ⅰ"))+"处，II级施工"+str(allLevels.count("Ⅱ"))+"处，III级施工"+str(allLevels.count("Ⅲ"))+"处，盯控x处。"
    #换个字体，加个边框
    for row in work_sheet.iter_rows(min_row=2, max_row=work_sheet.max_row):
        for cell in row:
            work_sheet[cell.coordinate].font = font
            work_sheet[cell.coordinate].border = border
    target_workbook.save('制作结果'+str(datetime.date.today())+'.xlsx')
    print("well done")
    pass


if __name__ == '__main__':
    #从word取数据
    allText = getText()
    #填到Excel里
    fillExcel()
    pass

